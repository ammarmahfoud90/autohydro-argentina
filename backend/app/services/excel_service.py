"""
Excel export service for AutoHydro Argentina.
Generates professional .xlsx workbooks with multiple sheets.
"""

from io import BytesIO
from datetime import datetime
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

# ── Color palette ──────────────────────────────────────────────────────────────

_NAVY   = "1A365D"
_BLUE   = "2563EB"
_LBLUE  = "DBEAFE"
_GRAY   = "6B7280"
_LGRAY  = "F3F4F6"
_GREEN  = "16A34A"
_RED    = "DC2626"
_ORANGE = "D97706"
_YELLOW = "FEF9C3"
_WHITE  = "FFFFFF"

_METHOD_NAMES = {
    "rational": "Método Racional",
    "modified_rational": "Método Racional Modificado",
    "scs_cn": "Método SCS-CN",
}
_RISK_LABELS = {
    "muy_bajo": "MUY BAJO",
    "bajo": "BAJO",
    "moderado": "MODERADO",
    "alto": "ALTO",
    "muy_alto": "MUY ALTO",
}
_INFRA_NAMES = {
    "alcantarilla_menor": "Alcantarilla menor",
    "alcantarilla_mayor": "Alcantarilla mayor",
    "puente_menor": "Puente menor",
    "puente_mayor": "Puente mayor",
    "canal_urbano": "Canal urbano",
    "canal_rural": "Canal rural",
    "defensa_costera": "Defensa costera",
}


# ── Style helpers ──────────────────────────────────────────────────────────────

def _font(bold=False, size=10, color=None, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, size=size,
                color=color or "000000", italic=italic)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _border_thin() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _apply_header_row(ws, row: int, values: list, bg: str = _NAVY, fg: str = _WHITE):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = _font(bold=True, size=10, color=fg)
        c.fill = _fill(bg)
        c.alignment = _align("center")
        c.border = _border_thin()


def _apply_data_row(ws, row: int, values: list, bg: Optional[str] = None,
                    bold_first: bool = True, number_cols: Optional[set] = None):
    number_cols = number_cols or set()
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = _border_thin()
        if bg:
            c.fill = _fill(bg)
        if col == 1 and bold_first:
            c.font = _font(bold=True, color=_NAVY)
            c.alignment = _align("left")
        else:
            c.font = _font()
            c.alignment = _align("right" if col in number_cols else "left")
        if isinstance(val, float):
            c.number_format = "0.000"


def _auto_col_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)


def _title_block(ws, row: int, title: str, subtitle: str = "") -> int:
    """Write a styled title block, return next row."""
    c = ws.cell(row=row, column=1, value="AutoHydro Argentina")
    c.font = _font(bold=True, size=14, color=_NAVY)
    row += 1
    c = ws.cell(row=row, column=1, value=title)
    c.font = _font(bold=True, size=12, color=_BLUE)
    row += 1
    if subtitle:
        c = ws.cell(row=row, column=1, value=subtitle)
        c.font = _font(size=9, color=_GRAY, italic=True)
        row += 1
    row += 1
    return row


# ── Sheet builders ─────────────────────────────────────────────────────────────

class ExcelReportGenerator:
    """Generate professional Excel workbook with hydrological calculation results."""

    def __init__(
        self,
        project_name: str = "Proyecto Hidrológico",
        location: str = "Argentina",
        engineer_name: str = "Ing. Ammar Mahfoud",
        client: Optional[str] = None,
    ):
        self.project_name = project_name
        self.location = location
        self.engineer_name = engineer_name
        self.client = client

    def generate(self, calculation_data: dict[str, Any]) -> BytesIO:
        wb = Workbook()
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)

        self._build_resumen(wb, calculation_data)
        self._build_datos_entrada(wb, calculation_data)
        self._build_calculos(wb, calculation_data)
        self._build_comparacion_metodos(wb, calculation_data)

        if calculation_data.get("cn_sensitivity"):
            self._build_sensibilidad_cn(wb, calculation_data)

        if calculation_data.get("hydrograph_times"):
            self._build_hidrograma(wb, calculation_data)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    # ── Sheet 1: Resumen ─────────────────────────────────────────────────────

    def _build_resumen(self, wb: Workbook, data: dict[str, Any]):
        ws = wb.create_sheet("Resumen")
        ws.sheet_view.showGridLines = False

        row = _title_block(ws, 1, "Resumen Ejecutivo",
                           f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        # Project info block
        _apply_header_row(ws, row, ["INFORMACIÓN DEL PROYECTO"], bg=_NAVY)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

        info_rows = [
            ("Proyecto", self.project_name),
            ("Localización", self.location),
            ("Ingeniero", self.engineer_name),
        ]
        if self.client:
            info_rows.append(("Comitente", self.client))
        info_rows.extend([
            ("Ciudad de referencia", data.get("city", "—")),
            ("Provincia", data.get("province", "—")),
            ("Fecha de cálculo", datetime.now().strftime("%d/%m/%Y")),
        ])

        for k, v in info_rows:
            _apply_data_row(ws, row, [k, v], bg=_LGRAY if row % 2 == 0 else None,
                            number_cols=set())
            row += 1

        row += 1

        # Main results block
        _apply_header_row(ws, row, ["RESULTADOS PRINCIPALES", "Valor", "Unidad"], bg=_BLUE)
        row += 1

        risk_lbl = _RISK_LABELS.get(data.get("risk_level", ""), "—")
        results = [
            ("Caudal pico de diseño (Q)", round(data.get("peak_flow_m3s", 0), 3), "m³/s"),
            ("Caudal específico (q)",     round(data.get("specific_flow_m3s_km2", 0), 4), "m³/s/km²"),
            ("Intensidad de diseño (i)",  round(data.get("intensity_mm_hr", 0), 1), "mm/hr"),
            ("Período de retorno (T)",    data.get("return_period", "—"), "años"),
            ("Tiempo de concentración",   round(data.get("tc_adopted_minutes", 0), 1), "min"),
            ("Método utilizado",          _METHOD_NAMES.get(data.get("method", ""), "—"), "—"),
            ("Nivel de riesgo",           risk_lbl, "—"),
            ("Infraestructura",           _INFRA_NAMES.get(data.get("infrastructure_type", ""), "—"), "—"),
        ]

        if data.get("cn") is not None:
            results.insert(5, ("Número de Curva (CN)", round(data["cn"], 1), "—"))
        if data.get("runoff_coeff") is not None:
            results.insert(5, ("Coeficiente de escorrentía (C)", round(data["runoff_coeff"], 2), "—"))

        for i, (label, val, unit) in enumerate(results):
            bg = _LGRAY if i % 2 == 0 else None
            # Highlight peak flow row
            if "Caudal pico" in label:
                bg = _LBLUE
            _apply_data_row(ws, row, [label, val, unit], bg=bg, number_cols={2})
            # Bold the peak flow value
            if "Caudal pico" in label:
                ws.cell(row=row, column=2).font = _font(bold=True, size=11, color=_NAVY)
            row += 1

        row += 1

        # Recommendations block
        recs = data.get("risk_recommendations", {})
        if recs:
            _apply_header_row(ws, row, ["RECOMENDACIONES"], bg=_NAVY)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            row += 1
            for key, label in [("general", "Situación"), ("action", "Acción"), ("verification", "Verificaciones")]:
                val = recs.get(key, "")
                if val:
                    c1 = ws.cell(row=row, column=1, value=label)
                    c1.font = _font(bold=True, color=_NAVY)
                    c1.alignment = _align("left")
                    c1.border = _border_thin()
                    c2 = ws.cell(row=row, column=2, value=val)
                    c2.font = _font()
                    c2.alignment = _align("left", wrap=True)
                    c2.border = _border_thin()
                    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
                    ws.row_dimensions[row].height = 30
                    row += 1

        row += 2
        footer = f"Generado con AutoHydro Argentina — {self.engineer_name}"
        c = ws.cell(row=row, column=1, value=footer)
        c.font = _font(size=8, color=_GRAY, italic=True)

        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 14

    # ── Sheet 2: Datos de Entrada ────────────────────────────────────────────

    def _build_datos_entrada(self, wb: Workbook, data: dict[str, Any]):
        ws = wb.create_sheet("Datos de Entrada")
        ws.sheet_view.showGridLines = False

        row = _title_block(ws, 1, "Datos de Entrada de la Cuenca")

        _apply_header_row(ws, row, ["PARÁMETROS DE LA CUENCA", "Valor", "Unidad"], bg=_NAVY)
        row += 1

        basin_rows = [
            ("Área de la cuenca (A)",        data.get("area_km2", "—"),  "km²"),
            ("Longitud del cauce principal (L)", data.get("length_km", "—"), "km"),
            ("Pendiente media (S)",           data.get("slope", "—"),    "m/m"),
            ("Período de retorno (T)",        data.get("return_period", "—"), "años"),
            ("Duración de tormenta (t)",      data.get("duration_min", "—"), "min"),
        ]
        for i, (label, val, unit) in enumerate(basin_rows):
            _apply_data_row(ws, row, [label, val, unit], bg=_LGRAY if i % 2 == 0 else None, number_cols={2})
            row += 1

        row += 1

        # IDF block
        _apply_header_row(ws, row, ["DATOS IDF", "Valor", "Fuente"], bg=_BLUE)
        row += 1
        idf_rows = [
            ("Ciudad de referencia", data.get("city", "—"), "—"),
            ("Intensidad de diseño (i)", round(data.get("intensity_mm_hr", 0), 2), data.get("idf_source", "—")),
            ("Fuente IDF", data.get("idf_source", "—"), "—"),
        ]
        for i, (label, val, src) in enumerate(idf_rows):
            _apply_data_row(ws, row, [label, val, src], bg=_LGRAY if i % 2 == 0 else None, number_cols={2})
            row += 1

        row += 1

        # Method-specific params
        method = data.get("method", "")
        _apply_header_row(ws, row, ["PARÁMETROS DEL MÉTODO", "Valor", "Notas"], bg=_NAVY)
        row += 1

        method_rows = [
            ("Método de cálculo", _METHOD_NAMES.get(method, method), "—"),
        ]
        if data.get("runoff_coeff") is not None:
            method_rows.append(("Coeficiente de escorrentía (C)", round(data["runoff_coeff"], 3), "—"))
        if data.get("cn") is not None:
            method_rows.extend([
                ("Número de Curva (CN)",  round(data["cn"], 1), "Calculado por usos de suelo"),
                ("Retención máxima (S)",  round(data.get("s_mm", 0), 2), "mm"),
                ("Abstracción inicial (Ia)", round(data.get("ia_mm", 0), 2), "mm"),
                ("Lámina escorrentía (Q)", round(data.get("runoff_depth_mm", 0), 2), "mm"),
            ])
        if data.get("areal_reduction_k") is not None:
            method_rows.append(("Factor reducción areal (K)", round(data["areal_reduction_k"], 4), "Témez"))

        for i, (label, val, note) in enumerate(method_rows):
            _apply_data_row(ws, row, [label, val, note], bg=_LGRAY if i % 2 == 0 else None, number_cols={2})
            row += 1

        _auto_col_width(ws)

    # ── Sheet 3: Cálculos ────────────────────────────────────────────────────

    def _build_calculos(self, wb: Workbook, data: dict[str, Any]):
        ws = wb.create_sheet("Cálculos")
        ws.sheet_view.showGridLines = False

        row = _title_block(ws, 1, "Procedimiento de Cálculo Paso a Paso")

        method = data.get("method", "")

        steps: list[tuple[str, str, str]] = [
            ("PASO 1", "Tiempo de Concentración (Tc)",
             f"Tc adoptado = {data.get('tc_adopted_hours', 0):.3f} hr  = {data.get('tc_adopted_minutes', 0):.1f} min\n"
             f"(promedio de {len(data.get('tc_results', []))} fórmulas seleccionadas)"),
            ("PASO 2", "Intensidad de Diseño (i)",
             f"Ciudad: {data.get('city', '—')}\n"
             f"T = {data.get('return_period', '—')} años, t = {data.get('duration_min', '—')} min\n"
             f"i = {data.get('intensity_mm_hr', 0):.2f} mm/hr\n"
             f"Fuente: {data.get('idf_source', '—')}"),
        ]

        if method in ("rational", "modified_rational"):
            C = data.get("runoff_coeff", "?")
            A = data.get("area_km2", "?")
            i = data.get("intensity_mm_hr", "?")
            Q = data.get("peak_flow_m3s", "?")
            if method == "rational":
                formula = f"Q = C × i × A / 3.6\nQ = {C} × {i:.2f} × {A} / 3.6 = {Q:.3f} m³/s"
            else:
                K = data.get("areal_reduction_k", 1.0)
                formula = (f"Q = K × C × i × A / 3.6\n"
                           f"K (Témez) = {K:.4f}\n"
                           f"Q = {K:.4f} × {C} × {i:.2f} × {A} / 3.6 = {Q:.3f} m³/s")
            steps.append(("PASO 3", "Cálculo del Caudal Pico (Método Racional)", formula))

        elif method == "scs_cn":
            CN = data.get("cn", "?")
            S = data.get("s_mm", "?")
            Ia = data.get("ia_mm", "?")
            Q_mm = data.get("runoff_depth_mm", "?")
            Qp = data.get("peak_flow_m3s", "?")
            tc = data.get("tc_adopted_hours", "?")
            steps.extend([
                ("PASO 3", "Cálculo del Número de Curva y Retención",
                 f"CN = {CN:.1f}\n"
                 f"S = 25400/CN - 254 = {S:.2f} mm\n"
                 f"Ia = 0.2 × S = {Ia:.2f} mm  (abstracción inicial)"),
                ("PASO 4", "Lámina de Escorrentía Directa (Ecuación SCS)",
                 f"P = intensidad × duración = precipitación total\n"
                 f"Q = (P - Ia)² / (P - Ia + S)\n"
                 f"Q_escorrentía = {Q_mm:.2f} mm"),
                ("PASO 5", "Caudal Pico (Hidrograma Unitario SCS)",
                 f"Tp = 0.6 × Tc = 0.6 × {tc:.3f} = {0.6*float(tc):.3f} hr\n"
                 f"Qp = 0.208 × A × Q_mm / Tp\n"
                 f"Qp = 0.208 × {data.get('area_km2', '?')} × {Q_mm:.2f} / {0.6*float(tc):.3f} = {Qp:.3f} m³/s"),
            ])

        _apply_header_row(ws, row, ["Paso", "Descripción", "Fórmula / Resultado"], bg=_NAVY)
        row += 1

        for i, (step, desc, formula) in enumerate(steps):
            bg = _LGRAY if i % 2 == 0 else None
            ws.cell(row=row, column=1, value=step).font = _font(bold=True, color=_BLUE)
            ws.cell(row=row, column=2, value=desc).font = _font(bold=True, color=_NAVY)
            c3 = ws.cell(row=row, column=3, value=formula)
            c3.alignment = _align("left", wrap=True)
            c3.font = _font(size=9)
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = _border_thin()
                if bg:
                    ws.cell(row=row, column=col).fill = _fill(bg)
            ws.row_dimensions[row].height = max(40, 15 * formula.count("\n") + 20)
            row += 1

        # Tc table
        tc_results = data.get("tc_results", [])
        if tc_results:
            row += 1
            _apply_header_row(ws, row, ["FÓRMULA Tc", "Tc (hr)", "Tc (min)", "Aplicabilidad"], bg=_BLUE)
            row += 1
            for i, tc in enumerate(tc_results):
                bg = _LGRAY if i % 2 == 0 else None
                _apply_data_row(ws, row,
                                [tc.get("formulaName", ""), round(tc.get("tcHours", 0), 3),
                                 round(tc.get("tcMinutes", 0), 1), tc.get("applicability", "")],
                                bg=bg, number_cols={2, 3})
                row += 1
            _apply_data_row(ws, row,
                            ["Tc Adoptado (promedio)", round(data.get("tc_adopted_hours", 0), 3),
                             round(data.get("tc_adopted_minutes", 0), 1), "Valor utilizado en el cálculo"],
                            bg=_LBLUE, number_cols={2, 3})
            for col in range(1, 5):
                ws.cell(row=row, column=col).font = _font(bold=True, color=_NAVY)

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 55

    # ── Sheet 4: Comparación de Métodos ──────────────────────────────────────

    def _build_comparacion_metodos(self, wb: Workbook, data: dict[str, Any]):
        ws = wb.create_sheet("Comparacion Metodos")
        ws.sheet_view.showGridLines = False

        row = _title_block(ws, 1, "Comparación de Métodos de Cálculo")

        method_comp = data.get("method_comparison", [])
        if not method_comp:
            ws.cell(row=row, column=1, value="No hay datos de comparación disponibles.")
            return

        _apply_header_row(ws, row, ["Método", "Q pico (m³/s)", "Tc (min)", "i (mm/hr)", "Notas"], bg=_NAVY)
        row += 1

        selected = data.get("method", "")
        chart_data_start = row
        for i, m in enumerate(method_comp):
            bg = _LBLUE if m.get("method") == selected else (_LGRAY if i % 2 == 0 else None)
            _apply_data_row(ws, row,
                            [m.get("methodName", ""), round(m.get("peakFlow", 0), 3),
                             round(m.get("tc", 0) * 60, 1), round(m.get("intensity", 0), 1),
                             m.get("notes", "")],
                            bg=bg, number_cols={2, 3, 4})
            if m.get("method") == selected:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).font = _font(bold=True, color=_NAVY)
            row += 1
        chart_data_end = row - 1

        # Bar chart
        if chart_data_end >= chart_data_start:
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.title = "Comparación de Caudales Pico"
            chart.y_axis.title = "Q (m³/s)"
            chart.x_axis.title = "Método"
            chart.style = 10
            chart.width = 16
            chart.height = 10

            data_ref = Reference(ws, min_col=2, min_row=chart_data_start - 1,
                                 max_row=chart_data_end)
            cats = Reference(ws, min_col=1, min_row=chart_data_start, max_row=chart_data_end)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, f"A{row + 2}")

        _auto_col_width(ws)

    # ── Sheet 5: Sensibilidad CN ──────────────────────────────────────────────

    def _build_sensibilidad_cn(self, wb: Workbook, data: dict[str, Any]):
        ws = wb.create_sheet("Sensibilidad CN")
        ws.sheet_view.showGridLines = False

        row = _title_block(ws, 1, "Análisis de Sensibilidad — Número de Curva (CN)",
                           "Evaluación del impacto de ±5 unidades en CN sobre el caudal pico")

        cn_sens = data.get("cn_sensitivity", [])

        _apply_header_row(ws, row, ["Escenario", "CN", "Q pico (m³/s)", "Variación (%)"], bg=_NAVY)
        row += 1
        chart_start = row

        for i, point in enumerate(cn_sens):
            label = point.get("label", "")
            is_base = label == "CN"
            bg = _LBLUE if is_base else (_LGRAY if i % 2 == 0 else None)
            var_pct = point.get("variation_pct", 0)
            var_str = "Base" if is_base else f"{'+' if var_pct >= 0 else ''}{var_pct:.1f}%"
            _apply_data_row(ws, row,
                            [label, round(point.get("cn", 0), 1),
                             round(point.get("peak_flow_m3s", 0), 3), var_str],
                            bg=bg, number_cols={2, 3})
            if is_base:
                for col in range(1, 5):
                    ws.cell(row=row, column=col).font = _font(bold=True, color=_NAVY)
            row += 1
        chart_end = row - 1

        # Bar chart
        if chart_end >= chart_start:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Sensibilidad del Caudal ante Variación de CN"
            chart.y_axis.title = "Q pico (m³/s)"
            chart.style = 10
            chart.width = 14
            chart.height = 9

            data_ref = Reference(ws, min_col=3, min_row=chart_start - 1, max_row=chart_end)
            cats = Reference(ws, min_col=1, min_row=chart_start, max_row=chart_end)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, f"A{row + 2}")

        _auto_col_width(ws)

    # ── Sheet 6: Hidrograma ───────────────────────────────────────────────────

    def _build_hidrograma(self, wb: Workbook, data: dict[str, Any]):
        ws = wb.create_sheet("Hidrograma")
        ws.sheet_view.showGridLines = False

        row = _title_block(ws, 1, "Hidrograma de Escorrentía — Método SCS",
                           "Basado en el Hidrograma Unitario Adimensional SCS (USDA-SCS, 1986)")

        times = data.get("hydrograph_times", [])
        flows = data.get("hydrograph_flows", [])
        Tp = data.get("time_to_peak_hr", 0)
        Tb = data.get("base_time_hr", 0)
        vol = data.get("runoff_volume_m3", 0)

        # Summary row
        _apply_header_row(ws, row, ["RESUMEN DEL HIDROGRAMA"], bg=_NAVY)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1

        Qp = max(flows) if flows else 0
        summary = [
            ("Caudal pico (Qp)", f"{Qp:.3f} m³/s"),
            ("Tiempo al pico (Tp)", f"{Tp:.2f} hr"),
            ("Tiempo base (Tb)", f"{Tb:.2f} hr"),
            ("Volumen de escorrentía", f"{vol:.0f} m³" if vol < 1000 else f"{vol/1000:.2f} × 10³ m³"),
        ]
        for i, (lbl, val) in enumerate(summary):
            _apply_data_row(ws, row, [lbl, val], bg=_LGRAY if i % 2 == 0 else None, number_cols=set())
            row += 1

        row += 1

        # Ordinate table
        _apply_header_row(ws, row, ["t (hr)", "t/Tp", "Q (m³/s)"], bg=_BLUE)
        row += 1
        data_start = row

        for i, (t, q) in enumerate(zip(times, flows)):
            t_tp = (t / Tp) if Tp and Tp > 0 else 0.0
            is_peak = abs(q - Qp) < 1e-9 and Qp > 0
            bg = _LBLUE if is_peak else (_LGRAY if i % 2 == 0 else None)
            _apply_data_row(ws, row,
                            [round(t, 3), round(t_tp, 3), round(q, 4)],
                            bg=bg, bold_first=False, number_cols={1, 2, 3})
            if is_peak:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).font = _font(bold=True, color=_NAVY)
            row += 1
        data_end = row - 1

        # Line chart
        if data_end > data_start:
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "standard"
            chart.title = "Hidrograma de Escorrentía"
            chart.y_axis.title = "Q (m³/s)"
            chart.x_axis.title = "Tiempo (hr)"
            chart.style = 10
            chart.width = 18
            chart.height = 11

            flow_ref = Reference(ws, min_col=3, min_row=data_start - 1, max_row=data_end)
            time_ref = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
            chart.add_data(flow_ref, titles_from_data=True)
            chart.set_categories(time_ref)
            ws.add_chart(chart, f"E{data_start}")

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 14
